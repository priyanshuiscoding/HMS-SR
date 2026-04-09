from __future__ import annotations

import json
import re
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_FILE = ROOT / "ALL COMPANY GOUDOWN STOCK 14-03-2026.xlsx"
OUTPUT_FILE = ROOT / "backend" / "src" / "data" / "godownInventory.generated.js"


def slugify(value: str) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return normalized or "item"


def clean_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).replace("\n", " ").strip()


def normalize_label(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def detect_header(row_values: list[str]) -> bool:
    labels = [normalize_label(value) for value in row_values if value]
    return any("name" in label or "medicine" in label for label in labels) and any(
        any(token in label for token in ("qty", "quantity", "amount", "remainingstock", "totalkg", "pack"))
        for label in labels
    )


def find_column(row_values: list[str], candidates: tuple[str, ...]) -> int | None:
    for index, value in enumerate(row_values):
        label = normalize_label(value)
        if not label:
            continue
        if any(candidate in label for candidate in candidates):
            return index
    return None


def parse_numeric_quantity(raw_value: str) -> tuple[float, str]:
    text = clean_text(raw_value).upper()
    if not text:
        return 0.0, ""

    preferred_patterns = [
        (r"=\s*(\d+(?:\.\d+)?)\s*(TAB|CAP|KG|GM|ML|LTR|LITRE|PACK|BOX|JAR|BOTTLE)?", 1),
        (r"(\d+(?:\.\d+)?)\s*(KG|GM|ML|LTR|LITRE|TAB|CAP|PACK|BOX|JAR|BOTTLE)", 1),
    ]

    for pattern, value_group in preferred_patterns:
        match = re.search(pattern, text)
        if match:
            quantity = float(match.group(value_group))
            unit = clean_text(match.group(value_group + 1)) if match.lastindex and match.lastindex > value_group else ""
            return quantity, unit

    numeric_matches = [float(item) for item in re.findall(r"\d+(?:\.\d+)?", text)]
    if numeric_matches:
        return max(numeric_matches), ""

    return 0.0, ""


def infer_formulation(name: str, packaging: str) -> str:
    combined = f"{name} {packaging}".lower()
    mapping = [
        ("oil", "Oil"),
        ("taila", "Oil"),
        ("arishta", "Liquid"),
        ("asava", "Liquid"),
        ("kwath", "Kwath"),
        ("syrup", "Syrup"),
        ("avaleh", "Avaleha"),
        ("leh", "Leham"),
        ("guggul", "Tablet"),
        ("vati", "Tablet"),
        ("tab", "Tablet"),
        ("cap", "Capsule"),
        ("capsule", "Capsule"),
        ("churna", "Churna"),
        ("powder", "Powder"),
        ("bhasma", "Bhasma"),
        ("ext", "Extract"),
        ("extract", "Extract"),
        ("cream", "Cream"),
        ("soap", "Soap"),
    ]
    for token, formulation in mapping:
        if token in combined:
            return formulation
    return "General"


def load_records():
    workbook = load_workbook(SOURCE_FILE, data_only=True)
    suppliers = OrderedDict()
    medicines = OrderedDict()
    batches = []
    transactions = []

    for sheet in workbook.worksheets:
        header_row_number = None
        header_values = None

        for row_index in range(1, min(sheet.max_row, 8) + 1):
            row_values = [clean_text(value) for value in next(sheet.iter_rows(min_row=row_index, max_row=row_index, values_only=True))]
            if detect_header(row_values):
                header_row_number = row_index
                header_values = row_values
                break

        if not header_row_number or not header_values:
            continue

        name_index = find_column(header_values, ("nameofmedicine", "medicine", "name"))
        company_index = find_column(header_values, ("company",))
        packaging_index = find_column(header_values, ("packaging", "packing", "packageqty"))
        quantity_index = find_column(header_values, ("remainingstock", "totalkg", "quantity", "qty", "quntity", "qyt", "amount", "noofpacks"))
        map_index = find_column(header_values, ("map", "drumno"))
        botanical_index = find_column(header_values, ("botanicalname",))

        if name_index is None or quantity_index is None:
            continue

        default_company = sheet.title.strip()
        supplier_id = f"sup-godown-{slugify(default_company)}"
        suppliers.setdefault(
            supplier_id,
            {
                "id": supplier_id,
                "name": default_company,
                "phone": "",
                "city": "",
                "isImported": True,
            },
        )

        for row_number, row in enumerate(
            sheet.iter_rows(min_row=header_row_number + 1, values_only=True),
            start=header_row_number + 1,
        ):
            values = [clean_text(value) for value in row]

            if not any(values):
                continue

            medicine_name = values[name_index] if name_index < len(values) else ""
            if not medicine_name:
                continue

            if normalize_label(medicine_name) in {"avaleh", "tablet", "capsule"}:
                continue

            company_name = (
                values[company_index] if company_index is not None and company_index < len(values) and values[company_index] else default_company
            )
            packaging = values[packaging_index] if packaging_index is not None and packaging_index < len(values) else ""
            quantity_text = values[quantity_index] if quantity_index < len(values) else ""
            location = values[map_index] if map_index is not None and map_index < len(values) else ""
            botanical_name = values[botanical_index] if botanical_index is not None and botanical_index < len(values) else ""

            if not quantity_text and not packaging:
                continue

            supplier_id = f"sup-godown-{slugify(company_name)}"
            suppliers.setdefault(
                supplier_id,
                {
                    "id": supplier_id,
                    "name": company_name,
                    "phone": "",
                    "city": "",
                    "isImported": True,
                },
            )

            medicine_key = slugify(f"{company_name}-{medicine_name}")
            medicine_id = f"med-godown-{medicine_key}"

            if medicine_id not in medicines:
                _, quantity_unit = parse_numeric_quantity(quantity_text)
                medicines[medicine_id] = {
                    "id": medicine_id,
                    "code": f"GDN-{len(medicines) + 1:04d}",
                    "name": medicine_name,
                    "formulation": infer_formulation(medicine_name, packaging),
                    "category": company_name,
                    "unit": quantity_unit.lower() or "unit",
                    "reorderLevel": 0,
                    "price": 0,
                    "gstPercent": 0,
                    "company": company_name,
                    "defaultPackaging": packaging,
                    "isImported": True,
                }

            parsed_quantity, quantity_unit = parse_numeric_quantity(quantity_text)
            batch_id = f"batch-godown-{slugify(sheet.title)}-{row_number}"
            batch_number = f"GODOWN-{slugify(sheet.title).upper()}-{row_number:03d}"

            batches.append(
                {
                    "id": batch_id,
                    "medicineId": medicine_id,
                    "medicineName": medicine_name,
                    "batchNumber": batch_number,
                    "supplierId": supplier_id,
                    "receivedDate": "2026-03-14",
                    "expiryDate": "",
                    "quantityReceived": parsed_quantity,
                    "quantityAvailable": parsed_quantity,
                    "purchasePrice": 0,
                    "sellingPrice": 0,
                    "packaging": packaging,
                    "rawQuantity": quantity_text,
                    "quantityUnit": quantity_unit.lower(),
                    "locationMap": location,
                    "company": company_name,
                    "sourceSheet": sheet.title,
                    "botanicalName": botanical_name,
                    "isImported": True,
                }
            )

            transactions.append(
                {
                    "id": f"txn-godown-{slugify(sheet.title)}-{row_number}",
                    "transactionDate": f"2026-03-14T08:{row_number % 60:02d}:00",
                    "medicineId": medicine_id,
                    "medicineName": medicine_name,
                    "batchId": batch_id,
                    "type": "receipt",
                    "quantity": parsed_quantity,
                    "referenceNumber": f"GODOWN-OPEN-{slugify(sheet.title).upper()}",
                    "note": f"Imported from godown sheet {sheet.title}",
                    "isImported": True,
                }
            )

    summary = {
        "importedOn": "2026-03-26",
        "sourceFile": SOURCE_FILE.name,
        "supplierCount": len(suppliers),
        "medicineCount": len(medicines),
        "batchCount": len(batches),
        "transactionCount": len(transactions),
    }

    return {
        "summary": summary,
        "suppliers": list(suppliers.values()),
        "medicineMasters": list(medicines.values()),
        "inventoryBatches": batches,
        "stockTransactions": transactions,
    }


def main():
    data = load_records()
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_FILE.write_text(
        "export const godownInventoryImport = " + json.dumps(data, indent=2) + ";\n",
        encoding="utf-8",
    )
    print(
        f"Wrote {data['summary']['medicineCount']} medicines and {data['summary']['batchCount']} batches to {OUTPUT_FILE}"
    )


if __name__ == "__main__":
    main()
